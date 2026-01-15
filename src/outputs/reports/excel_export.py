"""
Excel export functionaliteit.
"""
from pathlib import Path
from datetime import datetime
from typing import Optional
import logging

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError:
    pd = None
    Workbook = None

from ...database.connection import Database, get_connection
from ...database.queries import AccountQueries, MetricsQueries, PostQueries
from ...config.settings import COUNTRY_NAMES_NL, PLATFORM_NAMES_NL, EXPORTS_DIR

logger = logging.getLogger(__name__)


def export_monthly_report(
    year_month: str,
    output_path: Optional[Path] = None,
    db: Optional[Database] = None
) -> Path:
    """
    Exporteer maandelijks rapport naar Excel.
    """
    if pd is None or Workbook is None:
        raise ImportError("pandas en openpyxl zijn vereist voor Excel export")

    db = db or get_connection()

    if output_path is None:
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        output_path = EXPORTS_DIR / f"rapport_{year_month}.xlsx"

    # Create workbook
    wb = Workbook()

    # Sheet 1: Overzicht
    ws_overview = wb.active
    ws_overview.title = "Overzicht"
    _add_overview_sheet(ws_overview, year_month, db)

    # Sheet 2: Per Account
    ws_accounts = wb.create_sheet("Per Account")
    _add_accounts_sheet(ws_accounts, year_month, db)

    # Sheet 3: Per Platform
    ws_platforms = wb.create_sheet("Per Platform")
    _add_platforms_sheet(ws_platforms, year_month, db)

    # Sheet 4: Per Land
    ws_countries = wb.create_sheet("Per Land")
    _add_countries_sheet(ws_countries, year_month, db)

    # Save
    wb.save(output_path)
    logger.info(f"Excel rapport opgeslagen: {output_path}")

    return output_path


def _add_overview_sheet(ws, year_month: str, db):
    """Voeg overzicht sheet toe."""
    # Title
    ws['A1'] = f"NL Ambassade Social Media Rapport - {year_month}"
    ws['A1'].font = Font(size=16, bold=True)
    ws.merge_cells('A1:E1')

    ws['A3'] = f"Gegenereerd: {datetime.now().strftime('%Y-%m-%d %H:%M')}"

    # Get data
    metrics = MetricsQueries.get_all_for_month(year_month, db)
    accounts = AccountQueries.get_all(db)

    # Summary stats
    row = 5
    ws[f'A{row}'] = "Samenvatting"
    ws[f'A{row}'].font = Font(bold=True)

    row += 1
    ws[f'A{row}'] = "Totaal accounts:"
    ws[f'B{row}'] = len(accounts)

    row += 1
    ws[f'A{row}'] = "Accounts met data:"
    ws[f'B{row}'] = len(metrics)

    if metrics:
        row += 1
        ws[f'A{row}'] = "Totaal volgers:"
        ws[f'B{row}'] = sum(m.avg_followers or 0 for m in metrics)

        row += 1
        ws[f'A{row}'] = "Totaal posts:"
        ws[f'B{row}'] = sum(m.total_posts for m in metrics)

        row += 1
        ws[f'A{row}'] = "Totaal likes:"
        ws[f'B{row}'] = sum(m.total_likes for m in metrics)

        row += 1
        ws[f'A{row}'] = "Gem. engagement rate:"
        avg_eng = sum(m.avg_engagement_rate or 0 for m in metrics) / len(metrics)
        ws[f'B{row}'] = f"{avg_eng:.4f}%"

    # Auto-width columns
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20


def _add_accounts_sheet(ws, year_month: str, db):
    """Voeg accounts detail sheet toe."""
    # Headers
    headers = [
        "Land", "Platform", "Handle", "Volgers", "Volger Groei",
        "Posts", "Likes", "Comments", "Shares", "Engagement Rate", "Ranking"
    ]

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    metrics = MetricsQueries.get_all_for_month(year_month, db)
    accounts = {a.id: a for a in AccountQueries.get_all(db)}

    # Sort by engagement rate
    sorted_metrics = sorted(
        metrics,
        key=lambda m: m.avg_engagement_rate or 0,
        reverse=True
    )

    for row_num, m in enumerate(sorted_metrics, 2):
        account = accounts.get(m.account_id)
        if not account:
            continue

        ws.cell(row=row_num, column=1, value=COUNTRY_NAMES_NL.get(account.country, account.country))
        ws.cell(row=row_num, column=2, value=PLATFORM_NAMES_NL.get(account.platform, account.platform))
        ws.cell(row=row_num, column=3, value=account.handle)
        ws.cell(row=row_num, column=4, value=m.avg_followers)
        ws.cell(row=row_num, column=5, value=m.follower_growth)
        ws.cell(row=row_num, column=6, value=m.total_posts)
        ws.cell(row=row_num, column=7, value=m.total_likes)
        ws.cell(row=row_num, column=8, value=m.total_comments)
        ws.cell(row=row_num, column=9, value=m.total_shares)
        ws.cell(row=row_num, column=10, value=f"{m.avg_engagement_rate:.4f}%" if m.avg_engagement_rate else "")
        ws.cell(row=row_num, column=11, value=row_num - 1)  # Ranking

    # Auto-width
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = 15


def _add_platforms_sheet(ws, year_month: str, db):
    """Voeg platform vergelijking sheet toe."""
    from ...analysis.benchmarks import get_platform_comparison

    platform_data = get_platform_comparison(year_month, db)

    # Headers
    headers = ["Platform", "Accounts", "Totaal Volgers", "Posts", "Gem. Engagement"]

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for row_num, (platform, data) in enumerate(platform_data.items(), 2):
        ws.cell(row=row_num, column=1, value=PLATFORM_NAMES_NL.get(platform, platform))
        ws.cell(row=row_num, column=2, value=data["accounts"])
        ws.cell(row=row_num, column=3, value=data["total_followers"])
        ws.cell(row=row_num, column=4, value=data["total_posts"])
        ws.cell(row=row_num, column=5, value=f"{data['avg_engagement_rate']:.4f}%")


def _add_countries_sheet(ws, year_month: str, db):
    """Voeg landen vergelijking sheet toe."""
    from ...analysis.benchmarks import get_regional_comparison

    regional_data = get_regional_comparison(year_month, db)

    # Headers
    headers = ["Land", "Platforms", "Totaal Volgers", "Posts", "Gem. Engagement"]

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    # Data
    for row_num, (country, data) in enumerate(regional_data.items(), 2):
        ws.cell(row=row_num, column=1, value=data["display_name"])
        ws.cell(row=row_num, column=2, value=", ".join(data["platforms"]))
        ws.cell(row=row_num, column=3, value=data["total_followers"])
        ws.cell(row=row_num, column=4, value=data["total_posts"])
        ws.cell(row=row_num, column=5, value=f"{data['avg_engagement_rate']:.4f}%")


def export_yearly_report(
    year: int,
    output_path: Optional[Path] = None,
    db: Optional[Database] = None
) -> Path:
    """
    Exporteer jaarrapport naar Excel.
    """
    if pd is None or Workbook is None:
        raise ImportError("pandas en openpyxl zijn vereist voor Excel export")

    db = db or get_connection()

    if output_path is None:
        EXPORTS_DIR.mkdir(parents=True, exist_ok=True)
        output_path = EXPORTS_DIR / f"jaarrapport_{year}.xlsx"

    wb = Workbook()

    # Maandelijks overzicht
    ws = wb.active
    ws.title = "Maandelijks Overzicht"

    headers = ["Maand", "Accounts", "Totaal Volgers", "Posts", "Likes", "Gem. Engagement"]

    header_fill = PatternFill(start_color="1e3a5f", end_color="1e3a5f", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = header_fill
        cell.font = header_font

    for month in range(1, 13):
        year_month = f"{year:04d}-{month:02d}"
        metrics = MetricsQueries.get_all_for_month(year_month, db)

        row = month + 1
        ws.cell(row=row, column=1, value=year_month)
        ws.cell(row=row, column=2, value=len(metrics))
        ws.cell(row=row, column=3, value=sum(m.avg_followers or 0 for m in metrics))
        ws.cell(row=row, column=4, value=sum(m.total_posts for m in metrics))
        ws.cell(row=row, column=5, value=sum(m.total_likes for m in metrics))

        if metrics:
            avg_eng = sum(m.avg_engagement_rate or 0 for m in metrics) / len(metrics)
            ws.cell(row=row, column=6, value=f"{avg_eng:.4f}%")

    wb.save(output_path)
    logger.info(f"Jaarrapport opgeslagen: {output_path}")

    return output_path
